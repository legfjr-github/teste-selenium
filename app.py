import streamlit as st
import requests
from bs4 import BeautifulSoup
import time
import openpyxl
from pathlib import Path
import os

# --- CONFIGURAÇÃO DA AUTOMAÇÃO ---
URL_API_LIKE = "https://hentaifox.com/includes/add_like.php"
URL_API_FAP = "https://hentaifox.com/includes/add_fap.php"

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
    'Accept': '*/*',
    'Accept-Language': 'en-US,en;q=0.5',
    'Origin': 'https://hentaifox.com',
    'DNT': '1',
    'Connection': 'keep-alive',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'X-Requested-With': 'XMLHttpRequest'
}
# --- FIM DA CONFIGURAÇÃO ---


def run_automation(caminho_arquivo, total_repeticoes, cookie_string):
    """
    Executa o processo de automação usando requisições diretas e retorna o caminho do arquivo de progresso.
    """
    if not cookie_string:
        st.error("O Cookie é obrigatório para autenticar suas ações. Por favor, preencha o campo.")
        return None

    caminho_salva = str(caminho_arquivo).replace('.xlsx', '_PROGRESSO.xlsx')

    try:
        wb = openpyxl.load_workbook(caminho_arquivo)
        ws = wb.active
        if ws['C1'].value != 'Status':
            ws['C1'] = 'Status'
            wb.save(caminho_salva)
    except Exception as e:
        st.error(f"Não foi possível abrir o arquivo Excel: {e}")
        return None

    log_placeholder = st.empty()
    log_text = ""

    session = requests.Session()
    session.headers.update(HEADERS)
    session.headers['Cookie'] = cookie_string

    for i in range(total_repeticoes):
        log_text += f"--- INICIANDO REPETIÇÃO {i+1} DE {total_repeticoes} ---\n"
        log_placeholder.code(log_text)
        
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            link_cell, nome_pagina_cell, status_cell = row[0], row[1], row[2]
            
            gallery_link = link_cell.value
            nome_pagina = nome_pagina_cell.value

            if not gallery_link or not nome_pagina:
                continue

            log_text += f"\n[Linha {row_index}, Rep {i+1}] Processando: '{nome_pagina}'\n"
            log_placeholder.code(log_text)

            try:
                # PASSO 1: Visitar a página da galeria para obter o CSRF token
                log_text += "  - Buscando página para extrair token de segurança...\n"
                log_placeholder.code(log_text)
                response_page = session.get(gallery_link)
                response_page.raise_for_status()

                soup = BeautifulSoup(response_page.text, 'lxml')
                token_input = soup.find('input', {'name': 'X-Csrf-Token'})
                
                if not token_input or not token_input.has_attr('value'):
                    raise ValueError("Não foi possível encontrar o X-Csrf-Token na página.")
                
                csrf_token = token_input['value']
                log_text += f"  - Token encontrado: ...{csrf_token[-6:]}\n"
                log_placeholder.code(log_text)

                # PASSO 2: Preparar dados para a requisição POST
                gallery_id = gallery_link.strip('/').split('/')[-1]
                
                # O payload (dados a serem enviados) usa a chave 'gallery_id'
                # --- CORRIGIDO AQUI ---
                payload = {'gallery_id': gallery_id}
                
                dynamic_headers = {
                    'X-Csrf-Token': csrf_token,
                    'Referer': gallery_link
                }

                # PASSO 3: Enviar requisição de LIKE
                log_text += f"  - Enviando LIKE para galeria ID {gallery_id}...\n"
                log_placeholder.code(log_text)
                response_like = session.post(URL_API_LIKE, data=payload, headers=dynamic_headers)
                response_like.raise_for_status()
                log_text += "  - LIKE enviado com sucesso.\n"
                log_placeholder.code(log_text)
                time.sleep(1.5)

                # PASSO 4: Enviar requisição de FAP
                log_text += f"  - Enviando FAP para galeria ID {gallery_id}...\n"
                log_placeholder.code(log_text)
                response_fap = session.post(URL_API_FAP, data=payload, headers=dynamic_headers)
                response_fap.raise_for_status()
                log_text += "  - FAP enviado com sucesso.\n"
                log_placeholder.code(log_text)
                
                status_cell.value = f"Sucesso na repetição {i+1}"
                log_text += f"  -> SUCESSO para '{nome_pagina}'\n"
                log_placeholder.code(log_text)
                time.sleep(1.5)

            except requests.exceptions.HTTPError as e:
                erro_msg = f"Erro HTTP {e.response.status_code} na rep {i+1}"
                status_cell.value = erro_msg
                log_text += f"  -> ERRO: {erro_msg}\n"
                log_placeholder.code(log_text)
            except Exception as e:
                erro_msg = f"Erro inesperado na rep {i+1}: {e}"
                status_cell.value = erro_msg
                log_text += f"  -> ERRO: {erro_msg}\n"
                log_placeholder.code(log_text)
            
            finally:
                try: wb.save(caminho_salva)
                except Exception as e: log_text += f"  !!! AVISO: Falha ao salvar progresso: {e}\n"; log_placeholder.code(log_text)

    log_text += "\n--- Processo concluído ---\n"
    log_placeholder.code(log_text)
    return caminho_salva

# --- Interface do Streamlit (sem alterações) ---
st.set_page_config(page_title="Automatizador de Cliques", layout="wide")
st.title("Automatizador de Cliques para HentaiFox")
st.write("Faça o upload de uma planilha .xlsx com os links na primeira coluna e os nomes na segunda.")

cookie = st.text_input(
    "Cole o valor do seu Cookie aqui", 
    type="password",
    help="Abra as Ferramentas de Desenvolvedor (F12) > Aba Rede > Selecione uma requisição > Headers > Role até 'Request Headers' e copie TODO o valor do campo 'Cookie'."
)

uploaded_file = st.file_uploader("Escolha a planilha Excel", type="xlsx")
total_repeticoes = st.number_input("Digite o número de repetições por link", min_value=1, value=1, step=1)

if st.button("Iniciar Automação"):
    if uploaded_file is not None and cookie:
        path = Path.cwd() / uploaded_file.name
        with open(path, "wb") as f: f.write(uploaded_file.getbuffer())
        
        st.info("Iniciando o processo de automação... Por favor, aguarde.")
        caminho_resultado = run_automation(path, total_repeticoes, cookie)
        
        if caminho_resultado and os.path.exists(caminho_resultado):
            st.success("Automação concluída!")
            with open(caminho_resultado, "rb") as file:
                st.download_button(
                    label="Baixar Planilha com Resultados",
                    data=file,
                    file_name=os.path.basename(caminho_resultado),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            os.remove(path)
            os.remove(caminho_resultado)
        else:
            st.error("A automação falhou ou foi interrompida.")
            if os.path.exists(path): os.remove(path)
    else:
        st.warning("Por favor, faça o upload de um arquivo e preencha seu Cookie para iniciar.")
